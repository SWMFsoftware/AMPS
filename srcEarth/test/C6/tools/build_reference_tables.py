#!/usr/bin/env python3
"""Rebuild C6 machine-readable references from locally obtained source files.

This utility is intentionally separate from run_C6.py.  Test execution must use
stable checked-in CSVs and must not depend on network access, PDF extraction
versions, or spreadsheet libraries.  The builder records the exact assumptions
used when converting the three publications:

* Smart--Shea: pdftotext output containing the printed 5x30-degree Table 1.
* CARI-7: pdftotext -layout output of FAA report DOT/FAA/AM-19/4.  Appendix
  pages are fixed in that report edition.  Tiny printed negative cutoff values
  (-0.01 GV) are clamped to physical zero and explicitly marked in source text.
* Gerontidou: author-supplied XLSX with latitude in column A and longitudes
  0..360 in the next 25 columns.  The duplicated 360-degree column is omitted.

The source PDFs/XLSX are not downloaded automatically.  This prevents an
unnoticed website or document revision from changing a validation reference.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import re
from pathlib import Path
from typing import List, Tuple

HEADER = [
    "epoch_year", "latitude_deg", "longitude_deg_east",
    "altitude_km", "rc_effective_gv", "source",
]


def write_csv(path: Path, rows: List[Tuple[object, ...]], compressed: bool = False) -> None:
    if not compressed:
        with path.open("w", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(HEADER)
            writer.writerows(rows)
        return

    # Produce a deterministic gzip member.  gzip.open() records the current wall
    # clock in the member header, which makes two otherwise identical reference
    # rebuilds have different SHA-256 hashes.  The C6 archive is source-controlled,
    # so use mtime=0 and an empty embedded filename to make reproducibility checks
    # compare compressed bytes directly, not only the decompressed CSV content.
    with path.open("wb") as raw_stream:
        with gzip.GzipFile(
            filename="", fileobj=raw_stream, mode="wb", mtime=0
        ) as gzip_stream:
            with io.TextIOWrapper(gzip_stream, encoding="utf-8", newline="") as stream:
                writer = csv.writer(stream, lineterminator="\n")
                writer.writerow(HEADER)
                writer.writerows(rows)


def extract_smart_shea(text_path: Path, output_path: Path) -> None:
    lines = text_path.read_text(errors="replace").splitlines()
    rows: List[Tuple[object, ...]] = []
    in_table = False
    for line in lines:
        if line.startswith("Table 1: Effective Vertical Cutoff"):
            in_table = True
            continue
        if not in_table:
            continue
        match = re.match(r"^\s*(-?\d+)\s+((?:-?\d+\.\d+\s*){12})\s*$", line)
        if not match:
            continue
        latitude = int(match.group(1))
        values = [float(value) for value in match.group(2).split()]
        rows.extend(
            (2000, latitude, longitude, 20.0, cutoff, "Smart-Shea Table 1")
            for longitude, cutoff in zip(range(0, 360, 30), values)
        )
        if latitude == -90:
            break
    expected = 37 * 12
    if len(rows) != expected:
        raise RuntimeError("Smart--Shea extraction produced %d rows, expected %d" % (len(rows), expected))
    write_csv(output_path, rows)


def extract_cari(text_path: Path, output_path: Path) -> None:
    pages = text_path.read_text(errors="replace").split("\f")
    # Zero-based pdftotext page indices for the first appendix page of each epoch.
    starts = {1965: 26, 1980: 98, 1990: 170, 1995: 242, 2000: 314, 2010: 386}
    rows: List[Tuple[object, ...]] = []
    for year, start in starts.items():
        data = {}
        for page in pages[start:start + 72]:
            longitudes = None
            for line in page.splitlines():
                tokens = line.split()
                if len(tokens) == 20:
                    try:
                        integers = [int(token) for token in tokens]
                    except ValueError:
                        integers = []
                    if (len(integers) == 20 and 0 <= integers[0] <= 359 and
                            all(integers[i + 1] - integers[i] == 1 for i in range(19))):
                        longitudes = integers
                        continue
                if longitudes is None or len(tokens) != 21:
                    continue
                try:
                    latitude = int(tokens[0])
                    values = [float(token.replace("-.00", "-0.00")) for token in tokens[1:]]
                except ValueError:
                    continue
                if -89 <= latitude <= 89:
                    for longitude, raw_value in zip(longitudes, values):
                        cutoff = max(0.0, raw_value)
                        source = "CARI-7 FAA table"
                        if raw_value < 0.0:
                            source += " (negative interpolation artifact clamped to zero)"
                        data[(latitude, longitude)] = (cutoff, source)
        expected = 179 * 360
        if len(data) != expected:
            raise RuntimeError("CARI %d extraction produced %d rows, expected %d" % (year, len(data), expected))
        for (latitude, longitude), (cutoff, source) in sorted(data.items()):
            rows.append((year, latitude, longitude, 20.0, cutoff, source))
    write_csv(output_path, rows, compressed=True)


def extract_gerontidou(xlsx_path: Path, output_path: Path, year: int) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required only for XLSX reference conversion") from exc
    worksheet = load_workbook(xlsx_path, data_only=True, read_only=True).active

    # The author-supplied workbook stores the first longitude header as the
    # descriptive string "/Lon  0" while the remaining headers are numeric
    # values 15, 30, ..., 345.  Parse the trailing number rather than requiring
    # every cell to be a numeric Excel value.  This makes the converter faithful
    # to the distributed workbook instead of to a manually cleaned copy.
    longitudes = []
    for column in range(2, 26):
        raw = worksheet.cell(3, column).value
        if isinstance(raw, (int, float)):
            longitude = float(raw)
        else:
            match = re.search(r"[-+]?\d+(?:\.\d+)?\s*$", str(raw))
            if match is None:
                raise RuntimeError(
                    "cannot parse Gerontidou longitude header in column %d: %r"
                    % (column, raw)
                )
            longitude = float(match.group(0))
        longitudes.append(longitude)

    if longitudes != [float(value) for value in range(0, 360, 15)]:
        raise RuntimeError(
            "unexpected Gerontidou longitude header: %s" % longitudes
        )
    rows: List[Tuple[object, ...]] = []
    for row_index in range(4, 41):
        latitude = float(worksheet.cell(row_index, 1).value)
        for offset, longitude in enumerate(longitudes, start=2):
            cutoff = float(worksheet.cell(row_index, offset).value)
            rows.append((
                year, latitude, longitude, 20.0, max(0.0, cutoff),
                "Gerontidou et al. author-supplied %d grid" % year,
            ))
    expected = 37 * 24
    if len(rows) != expected:
        raise RuntimeError("Gerontidou extraction produced %d rows, expected %d" % (len(rows), expected))
    write_csv(output_path, rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--smart-shea-text", type=Path)
    parser.add_argument("--cari-text", type=Path)
    parser.add_argument("--gerontidou-xlsx", type=Path)
    parser.add_argument("--gerontidou-year", type=int, default=2015)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    if args.smart_shea_text:
        extract_smart_shea(args.smart_shea_text, args.output_dir / "reference_C6_smart_shea_2000.csv")
    if args.cari_text:
        extract_cari(args.cari_text, args.output_dir / "reference_C6_cari7_1965_2010.csv.gz")
    if args.gerontidou_xlsx:
        extract_gerontidou(
            args.gerontidou_xlsx,
            args.output_dir / ("reference_C6_gerontidou_%d.csv" % args.gerontidou_year),
            args.gerontidou_year,
        )
    if not any((args.smart_shea_text, args.cari_text, args.gerontidou_xlsx)):
        parser.error("select at least one source file")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
